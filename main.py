import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from role_engine import score_job


def priority_level(score):
    if score >= 70:
        return "High Priority"
    elif score >= 40:
        return "Medium Priority"
    else:
        return "Low Priority"


def format_excel_file(filename):
    wb = load_workbook(filename)
    ws = wb.active

    for cell in ws[1]:
        cell.font = Font(bold=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            value = str(cell.value) if cell.value is not None else ""
            max_length = max(max_length, len(value))

        ws.column_dimensions[column_letter].width = min(max_length + 2, 55)

    link_col = None
    for cell in ws[1]:
        if cell.value == "Link":
            link_col = cell.column
            break

    if link_col:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=link_col)
            if cell.value:
                cell.hyperlink = cell.value
                cell.style = "Hyperlink"

    priority_col = None
    for cell in ws[1]:
        if cell.value == "Priority":
            priority_col = cell.column
            break

    if priority_col:
        for row in ws.iter_rows(min_row=2):
            priority = row[priority_col - 1].value

            if priority == "High Priority":
                fill = PatternFill(
                    start_color="C6EFCE",
                    end_color="C6EFCE",
                    fill_type="solid"
                )
            elif priority == "Medium Priority":
                fill = PatternFill(
                    start_color="FFF2CC",
                    end_color="FFF2CC",
                    fill_type="solid"
                )
            else:
                fill = PatternFill(
                    start_color="F4CCCC",
                    end_color="F4CCCC",
                    fill_type="solid"
                )

            for cell in row:
                cell.fill = fill

    wb.save(filename)


def main():
    jobs = pd.read_csv("saved_jobs.csv")
    jobs = jobs.fillna("")
    jobs.columns = jobs.columns.str.strip().str.lower()

    required_columns = ["title", "company", "location", "description", "link"]

    for col in required_columns:
        if col not in jobs.columns:
            raise ValueError(
                f"Missing column: {col}. Found columns: {list(jobs.columns)}"
            )

    results = []

    for _, job in jobs.iterrows():
        job_dict = {
            "title": str(job.get("title", "")),
            "company": str(job.get("company", "")),
            "location": str(job.get("location", "")),
            "description": str(job.get("description", "")),
            "link": str(job.get("link", "")),
            "source": str(job.get("source", ""))
}

        role_result = score_job(job_dict)
        matched_role_groups = ", ".join(
            role_result["matched_role_groups"].keys()
        )

        matched_tech_keywords = ", ".join(
            role_result["matched_tech_keywords"]
        )

        matched_titles = ", ".join(
            role_result["matched_titles"]
        )

        results.append({
            "Title": job["title"],
            "Company": job["company"],
            "Location": job["location"],
            "Source": job.get("source", ""),
            "Date Found": job.get("date_found", ""),
            "Match Score": role_result["score"],
            "Priority": priority_level(role_result["score"]),
            "Primary Role": role_result["primary_role"],
            "Entry Level": role_result["entry_level"],
            "Senior Level": role_result["senior_level"],
            "Matched Role Groups": matched_role_groups,
            "Matched Tech Keywords": matched_tech_keywords,
            "Matched Titles": matched_titles,
            "Application Status": "Not Applied",
            "Notes": "",
            "Link": job["link"]
        })

    df = pd.DataFrame(results)
    df = df.sort_values(by=["Match Score"], ascending=False)

    df.to_csv("matched_jobs_v3.csv", index=False)
    df.to_excel("matched_jobs_v3.xlsx", index=False)

    format_excel_file("matched_jobs_v3.xlsx")

    print("V3 role-engine job matching complete.")
    print("Created:")
    print("- matched_jobs_v3.csv")
    print("- matched_jobs_v3.xlsx")


if __name__ == "__main__":
    main()